import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

url = "https://www.mercadolivre.com.br/mouse-sem-fio-recarregavel-wireless-optico-led-rgb-ergonmico-sem-fio-para-pc-desktop-notebook-tablet-smartphone-homeroarte-preto/p/MLB64093163#polycard_client=search-desktop&search_layout=grid&position=6&type=product&tracking_id=47bc0645-68d5-4f6e-a9bd-bfe71f0b45b5&wid=MLB4057383363&sid=search"

headers = {"User-Agent": "Mozilla/5.0"}
response = requests.get(url, headers=headers)

soup = BeautifulSoup(response.text, "html.parser")

preco = soup.select_one(".ui-pdp-price__second-line .andes-money-amount__fraction")

centavos = soup.select_one(".ui-pdp-price__second-line .andes-money-amount__cents")

valor = f"{preco.text}.{centavos.text if centavos else "00"}"

valor_float = float(valor)

#------------------------------------------------------------------------------------------------------------

def salvar_historico(preco):
    data = datetime.now().strftime("%d/%m/%Y %H:%M")

    if os.path.exists("historico_precos.csv"):
        df_existente = pd.read_csv("historico_precos.csv")

        if not df_existente.empty:
            ultimo_preco = df_existente.iloc[-1]["Preço"]
        else:
            ultimo_preco = None
    else:
        df_existente = None
        ultimo_preco = None

    alerta = False
    if ultimo_preco is not None and preco < ultimo_preco:
        alerta = True

    mudou = ultimo_preco != preco

    novo_dado = pd.DataFrame([{
    "Data": data,
    "Preço": preco,
    "Mudou": mudou,
    "Alerta": alerta
    }])

    if df_existente is not None:
        df = pd.concat([df_existente, novo_dado], ignore_index=True)
    else:
        df = novo_dado

    df.to_csv("historico_precos.csv", index=False)

    if mudou:
        print("Preço mudou! Registro atualizado.")
    else:
        print("Preço não mudou, mas registrado.")

    if alerta:
        print("ALERTA: O PREÇO CAIU!!!")
        print(f"De R$ {ultimo_preco} para R$ {preco}")

salvar_historico(valor_float)

#-------------------------------------------------------------------------------------------------------

df = pd.read_csv("historico_precos.csv")

df["Mudou"] = df["Mudou"].map({True: "Sim", False: "Não"})
df["Alerta"] = df["Alerta"].map({True: "Sim", False: "Não"})

with pd.ExcelWriter("historico_precos.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Historico")
    worksheet = writer.sheets["Historico"]

    cor = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    fonte = Font(color="FFFFFF", bold=True)

    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = cor
        cell.font = fonte
        cell.alignment = Alignment(horizontal="center")

    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        worksheet.column_dimensions[col_letter].width = max_length + 5
    
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = "A2"

    col_preco = 2
    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=col_preco).number_format = "R$ #,##0.00"

    for row in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row, column=1).number_format = "DD/MM/YYYY HH:MM"

    cor_alerta = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(2, worksheet.max_row + 1):
        alerta = worksheet.cell(row=row, column=4).value

        if alerta == "Sim":
            for col in range (1, worksheet.max_column + 1):
                worksheet.cell(row=row, column=col).fill = cor_alerta

print("Relatório Excel gerado com sucesso!")


